# intern modules
import re

#external modules
from openpyxl import Workbook, load_workbook


# django modules
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect
from django.db.utils import DatabaseError, OperationalError, ProgrammingError
from django.core.cache import cache


# models
from assets_users.models import Assets
from assets_users.models import Departments
from assets_users.models import Users

# local modules

# users modules
from .local_modules.users.validations import validate_data
from .local_modules.users.hashing import hashing_password, compare_password
from .local_modules.users.update_user import update_user
from .local_modules.users.registration_users import registration 
from .local_modules.users.login_user import login, get_password, get_data_user, set_sessions_user

# department modules
from .local_modules.department_list import department_list

# database_excpetions modules
from .local_modules.database_exceptions import exception_db_response

# assets modules
from .local_modules.assets.register_asset import registration_asset, validate_dni
from .local_modules.assets.styling_excel import styling_assets_excel


# vista para registrar usuario
def user_register(request):
    
    context = {} 
    try:
        if request.method == "POST":
            post_data = { 
                'id_card':request.POST['identification_card'],
                'name':request.POST['name'], 'surnames':request.POST['surnames'],
                'phone':request.POST['phone'],'department':request.POST['department'],
                'age':request.POST['age'],'email':request.POST['email'].replace(".com", ".oo")
            }

            password_hash = hashing_password(request.POST["password"]) 

            if validate_data(post_data):  
                if registration(post_data, password_hash, department_list()):
                    context['response'] = '1'
                    return render(request, 'registration.html', context)
                else: 
                    context['response'] = '2'
                    context['departments'] = department_list()
                    context['data'] = post_data
                    return render(request, 'registration.html', context)
            else: 
                context['response'] = '0'  # contexto que se envía al html y se rescata para JS
                context['departments'] = department_list()
                context['data'] = post_data
                return render(request, 'registration.html', context)

    except DatabaseError as e:

                exception_db_response(
                '3', 
                department_list(), 
                render(request, 'registration.html', context))
                
    except OperationalError as e:

                exception_db_response(
                '4', 
                department_list(), 
                render(request, 'registration.html', context))
    except ProgrammingError as e:

                exception_db_response(
                '5', 
                department_list(), 
                render(request, 'registration.html', context))

# vista para validar usuario
def user_login(request):

    context = {}
    if request.method == 'POST':
            
        post_data = {'identification_card': request.POST['identification_card'],
                        'password': request.POST['password']}

        object_request = login(post_data['identification_card'])
        password_user = get_password(object_request) 

        if object_request and password_user: # si ha encontrado algo con dicha cedula
            
                if compare_password(post_data['password'], password_user):
                    try:
                        data_user = get_data_user(object_request)
                        
                        administrator_auth = bool()

                        for data in object_request:
                            administrator_auth = data.administrator
                            break                

                        if not(administrator_auth):
                            set_sessions_user(request, data_user)
                            return redirect('/boxder/')
                        else:
                            set_sessions_user(request, data_user)
                            return redirect('/administrator/')    
                    except:
                        context['response'] = '0'
                        return render(request, 'index.html', context)
                else:
                    context['response'] = '1'
                    return render(request, 'index.html', context)

        else:
            context['response'] = '1'
            return render(request, 'index.html', context)

# vista para formulario
def registration_form(request):

    context = {'departments': department_list()}
    return render(request, 'registration.html', context)

# vista para registrar activos
def asset_registration(request):
    context = {}
    if request.method == "POST":
        data_asset = {'brand': request.POST['brand'],'model': request.POST['model'],
                      'cost': request.POST['cost'], 'weight': request.POST['weight'],
                      'provider': request.POST['provider'], 'util_life': request.POST['util_life'],
                      'name':request.POST['name'], 'user_id':request.POST['user_dni']}
        

        if validate_dni(data_asset['user_id']):
            
            if registration_asset(data_asset) == '1':

                context['response'] = '1'
            
            elif registration_asset(data_asset) == '2':
                
                context['response'] = '2'
            
            elif registration_asset(data_asset) == '3':

                context['response'] = '3'
        else:
            context['response'] = '4'
                
        return render(request, 'boxderindex.html', context)
        
# vista para página principal de la app
def boxder_index(request):

    context = {}
    try: 
        if request.session['name'] or request.session['surnames']:
                all_assets = Assets.objects.all();
                context['total_assets'] = all_assets.count()
                context['active_assets'] = Assets.objects.filter(actual_status="Activo").count()
                context['suspend_assets'] = Assets.objects.filter(actual_status="Suspendido").count()
                context['low_assets'] = Assets.objects.filter(actual_status="De baja").count()
                context['assets'] = Assets.objects.filter(user_id_id=request.session['id'])
                return render(request, 'boxderindex.html', context)

    except KeyError as e:
        return redirect('/inicio/')
  
# vista para página principal de administrador
def boxder_admin(request):


    context = {}
    try: 
        if request.session['name'] or request.session['surnames']:
                all_assets = Assets.objects.all();
                context['assets'] = all_assets
                context['total_assets'] = all_assets.count()
                print(context['total_assets'])
                return render(request, 'boxderAdmin.html', context)

    except KeyError as e:
        return redirect('/inicio/')

def configuration(request):
    context = {}
    context['departments'] = department_list()
    context['users'] = Users.objects.filter(identification_card=request.session['id'])
    return render(request, 'configuration_user.html', context)

def configuration_update(request):

    context = {}
    if request.method == 'POST':
        
        data = {"id": request.POST['identification_card'],
                "name": request.POST['name'],
                "surnames": request.POST['surnames'],
                "phone": request.POST['phone'],
                "department": request.POST['department'],
                "age": request.POST['age'],
                "email": request.POST['email']}

        if update_user(data) == '1':
            context['response'] = '1'

        elif update_user(data) == '2':
            context['response'] = '2'

        elif update_user(data) == '3':
            context['response'] = '3'
        
        context['users'] = Users.objects.filter(identification_card=request.session['id'])
        context['departments'] = department_list()
        return render(request, 'configuration_user.html', context)

def export_excel(request):

    wb = Workbook()
    sheet = wb.active
    
    activos = Assets.objects.all()

    datos = [("Code", "Brand", "Model", "Useful_life", "Cost", "Weight", 
            "admission_date", "actual_status", "provider", "user_id", "name")]
            
    for data in activos:
        assets_data = (data.code, data.brand, data.model,
                            data.useful_life, data.cost, data.weight,
                            data.admission_date, data.actual_status,
                            data.provider, data.user_id_id, data.name)
        
        datos.append(assets_data)
    
    sheet.merge_cells("A1:K1")
    
    styling_assets_excel(sheet, datos)
        
    response = HttpResponse(content_type='application/ms-excel')
    # response = tipo de respuesta que será un archivo de microsoft excel
    response['Content-Disposition'] = 'attachment; filename="reporteActivos.xlsx"'
    # añade el nombre del archivo 

    wb.save(response)

    return response # retorna la respuesta (el archivo excel)
    
    return HttpResponse("1")
            

        
