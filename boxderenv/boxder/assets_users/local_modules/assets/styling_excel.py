from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

def styling_assets_excel(sheet, datos):

    sheet['A1'].font = Font(color="FFFFFF", size=14)
    sheet.row_dimensions[1].height = 30
    sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
    sheet['A1'].fill = PatternFill(start_color="1173C9", end_color="1173C9", fill_type="solid")
    sheet["A1"] = "Activos"

    for row in datos:
        sheet.append(row)

    sheet.row_dimensions[2].height = 30 # accede a la fila 2 y coloca una altura de 30 px
    
    for row in sheet['A2':'K2']: # muestra todas las celdas entre la celda A2 y K2, o sea la rila 2
        for cell in row: # accede a cada celda en dicha fila
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row in sheet['A2':'K2']: # modificación de width a las columnas
        for col in row:
            if col.column == 4 or col.column == 6 or col.column == 9:
                sheet.column_dimensions[get_column_letter(col.column)].width = 10
            else:
                sheet.column_dimensions[get_column_letter(col.column)].width = 15
            
            col.font = Font(color="FFFFFF")
            col.fill = PatternFill(start_color="F6694B", end_color="1173C9", fill_type="solid")

    for row in sheet['A':'K']:
        for col in row:
            col.alignment = Alignment(horizontal="center", vertical="center")
         #   if col.row != 1 or col.row != 2:
         #       if col.column % 2 == 0:
         #           col.fill = PatternFill(start_color="F3E3DF", end_color="F3E3DF", fill_type="solid")
         #       else:
         #           col.fill = PatternFill(start_color="E6A595", end_color="E6A595", fill_type="solid")

    